import openpyxl
from openpyxl.styles import Alignment
from database.funcs import get_weather
from serializers.database_filters import GetWeatherFilters

from loguru import logger


async def create_xlsx(filename: str = 'weather_data.xlsx', limit: int = 10):
    if not filename.endswith('.xlsx'):
        logger.error('Invalid filename! Filename must end with ".xlsx"')
        return

    # Запрашиваем данные из базы данных
    weather_data = await get_weather(filters=GetWeatherFilters(limit=limit))

    # Создаем новый Excel-файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Weather Data"

    # Заголовки столбцов
    headers = ["Date Time", "Temperature", "Wind Speed", "Wind Direction", "Air Pressure", "Precipitation Type",
               "Precipitation Count"]

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)
        sheet.cell(row=1, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

    # Записываем данные в таблицу
    for row_num, data in enumerate(weather_data, 2):
        sheet.cell(row=row_num, column=1, value=data.datetime)
        sheet.cell(row=row_num, column=2, value=data.temperature)
        sheet.cell(row=row_num, column=3, value=data.wind_speed)

        # Переводим направление ветра в текстовый формат
        wind_direction_text = ""
        if 0 <= data.wind_direction < 22.5 or 337.5 <= data.wind_direction <= 360:
            wind_direction_text = "С"
        elif 22.5 <= data.wind_direction < 67.5:
            wind_direction_text = "СВ"
        elif 67.5 <= data.wind_direction < 112.5:
            wind_direction_text = "В"
        elif 112.5 <= data.wind_direction < 157.5:
            wind_direction_text = "ЮВ"
        elif 157.5 <= data.wind_direction < 202.5:
            wind_direction_text = "Ю"
        elif 202.5 <= data.wind_direction < 247.5:
            wind_direction_text = "ЮЗ"
        elif 247.5 <= data.wind_direction < 292.5:
            wind_direction_text = "З"
        elif 292.5 <= data.wind_direction < 337.5:
            wind_direction_text = "СЗ"

        sheet.cell(row=row_num, column=4, value=wind_direction_text)

        sheet.cell(row=row_num, column=5, value=data.air_pressure)
        sheet.cell(row=row_num, column=6, value=data.precipitation_type)
        sheet.cell(row=row_num, column=7, value=data.precipitation_count)

    # Сохраняем файл
    workbook.save(filename)
    logger.info(f"File {filename} was successfully saved!")

if __name__ == "__main__":
    import asyncio

    asyncio.run(create_xlsx())
